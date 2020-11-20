'''Create the Kids In Custody Report with Active Placement and without Active Placement'''
import pandas as pd, pyautogui as p, json, numpy as np, os, datetime as dt, stuff
from simple_salesforce import Salesforce, format_soql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def kic_rpt():
    '''Run the current kids in custody with and without placement report'''
    def column_size(sheet):
        '''Dynamically adjust the column sizes in excel sheet'''
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))+5
                else:
                    column_widths += [len(str(cell.value))+5]
        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i+1)].width = column_width		

    def soql_df(soql_query):
        '''Create a DF using SOQL query and normalizing JSON'''
        info = sf.query_all(soql_query)
        df = pd.json_normalize(info['records'])
        cols = [c for c in df.columns if 'attribute' not in c]
        cols = [c for c in cols if not c.endswith('__r')]
        df = df[cols].copy()
        return df

    def cus_rank(x):
        '''Rank anything that is not DSCYF/DFS, DFS, DSCYF as 1 for sorting'''
        if x in ('DSCYF/DFS', 'DFS', 'DSCYF'):
            return 2
        else:
            return 1

    def rank_null(x):
        '''Rank null dates above a normal date for sorting'''
        if x == None:
            return 1
        else:
            return 2

    #Sign into Salesforce.
    username = stuff.username
    #password = p.password('Enter your password', title='Salesforce Password')
    password = stuff.password
    orgid = ''
    url = ''
    sf = Salesforce(username = username, password = password, instance_url = url, organizationId = orgid)

    #Create the initial Kids in Custody query.
    q = f'''
    SELECT Custodian_Name__c,Start_Date__c,End_Date__c,Client_Name__r.DEL_PID__c,Client_Name__r.Name,Client_Name__r.Birthdate,
    Client_Name__r.DEL_Age__c, Client_Name__r.DEL_Age_Years__c,Client_Name__r.DEL_Custody_Start_Date__c,
    Client_Name__r.DEL_Runaway_Alert_Flag__c, Client_Name__r.DEL_MCI__c
    FROM DEL_Custody__c

    ORDER BY Client_Name__r.DEL_PID__c, Start_Date__c DESC
    '''

    df = soql_df(q)

    #Rank the custodian and null dates
    df['Custodian_Rank'] = df['Custodian_Name__c'].apply(cus_rank)
    df['Date_Rank'] = df['End_Date__c'].apply(rank_null)

    #Sort the data to get the most recent start date with a non DFS custodian first and open end date first as well.
    df = df.sort_values(by=['Client_Name__r.DEL_PID__c', 'Start_Date__c',
                              'Custodian_Rank', 'Date_Rank'], ascending = [True, False,True, True]).copy()

    #Drop duplicate by PID and then keep the ones that are DFS related with an open end date.
    df = df.drop_duplicates('Client_Name__r.DEL_PID__c').copy()

    df = df[df['Custodian_Name__c'].isin(['DSCYF/DFS', 'DFS', 'DSCYF'])].copy()

    df = df[df['End_Date__c'].isnull()].copy()

    #Drop the rank columns and date columns
    df = df.drop(['Start_Date__c','End_Date__c','Custodian_Rank', 'Date_Rank'], axis = 'columns').copy()

    #Change Age_Years into a numeric.
    df = df.astype({'Client_Name__r.DEL_Age_Years__c' : 'int32'}).copy()

    #Make a list of all the PIDs
    pids = list(df['Client_Name__r.DEL_PID__c'])

    #Query for Eligibility Information
    q1 =format_soql('''
    SELECT  Person_LkID__r.DEL_PID__c,Start_Date__c, End_Date__c,  Aid_Category_Description__c
    FROM DEL_Income_Eligibility__c
    WHERE Person_LkID__r.DEL_PID__c IN {pids}
    AND Eligibility_Income_Type__c = 'Medicaid Eligibility'
    ORDER BY End_Date__c DESC
    ''', pids = pids)

    #Create Dataframe
    #JSON Normalize will remove the nested dictionaries
    #and then we will keep the columns that do not have the word 'attribute in it'. Then remove columns that end with __r.
    df1 = soql_df(q1)

    #Sort data by End_Date__c and remove duplicates by PID
    df1 = df1.sort_values(by = ['Person_LkID__r.DEL_PID__c','End_Date__c'], ascending = [True, False]).copy()
    df1 = df1.drop_duplicates('Person_LkID__r.DEL_PID__c').copy()

    #Merge df1 and df and then drop Person_LkID__r.DEL_PID__c.
    df1 =  df.merge(df1, how='left',left_on = 'Client_Name__r.DEL_PID__c', right_on = 'Person_LkID__r.DEL_PID__c').copy()
    df1 = df1.drop(columns=['Person_LkID__r.DEL_PID__c']).copy()

    #Create the SOQL query for all the Placements for those PIDs.
    q2=format_soql('''
    SELECT PID__c, Name,Division__c, Service_Name__c,
        Case_Number__r.Assigned_Worker__r.Name , Case_Number__r.Assigned_Worker__r.DEL_Service_Area__c,
        Case_Number__r.Assigned_Supervisor__r.Name, State__c, Placement_Start_Date_Division_Wide__c,
        Case_Number__r.Name, Case_Type__c, Placement_Type_Formula__c
    FROM DEL_Placement__c
    WHERE PID__c IN {pids}
    AND Service_Name__c != null
    AND Placement_Start_Date_Division_Wide__c != null
    AND Placement_End_Date_Division_Wide__c = null
    AND Latest_Version_Placement__c = True
    ''', pids = pids)

    #Run the query and put it into the Pandas DataFrame.
    df2 = soql_df(q2)

    #Merge df1 with df2 and rename columns and order them correctly.
    df3 = df1.merge(df2, how='left',left_on = 'Client_Name__r.DEL_PID__c', right_on = 'PID__c').copy()
    df3 = df3.drop(columns = ['PID__c']).copy()

    columns = [
        'Who Has Custody','PID', 'Name', 'Birth Date', 'Age', 'Age (Years)', 'Custody Start Date', 'Runaway',
        'MCI', 'MA Start Date', 'MA End Date', 'MA Code',  'Placement ID',
        'Division', 'Service', 'State', 'Placement Start Date', 'Case Type','Placement Type', 'Case Owner',
        'Service Area', 'Supervisor', 'Case Number'
    ]

    df3.columns = columns #Rename columns

    ord_columns = [
        'PID', 'Name', 'Birth Date', 'Age', 'Age (Years)', 'Custody Start Date','Division','Service','Case Number', 'Case Type', 'Case Owner',
        'Supervisor','Service Area','State','Placement ID','Placement Start Date','Who Has Custody','Runaway', 'MCI','MA Start Date', 'MA End Date', 'MA Code',
        'Placement Type'
    ]

    df3 = df3[ord_columns].copy() #Order the columns

    #Kids in Custody with Active Placement
    wplacement = df3.loc[~df3['Placement ID'].isnull()].copy()
    wplacement.loc[wplacement['MA End Date']=='2299-12-31', 'MA End Date'] = '2050-12-31'
    wplacement.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date', 'MA Start Date', 'MA End Date']]= wplacement.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date', 'MA Start Date', 'MA End Date']].apply(pd.to_datetime) 
    wplacement1 = wplacement.copy() #For KIC with current DFS Placement Report

    #Drop additional column of Placement Type not needed for the report but needed for the other report mentioned above.
    wplacement = wplacement.drop('Placement Type', axis = 'columns').copy()

    #Kids in Custody without Active Placement
    woplacement = df3.loc[df3['Placement ID'].isnull()].copy()
    woppids = list(woplacement['PID']) #get pids for without placement.

    woq = format_soql('''
    SELECT PID__c, Name,Division__c, Service_Name__c, End_Reason__c,
        Case_Number__r.Assigned_Worker__r.Name , Case_Number__r.Assigned_Worker__r.DEL_Service_Area__c,
        Case_Number__r.Assigned_Supervisor__r.Name, State__c, Placement_Start_Date_Division_Wide__c,
        Placement_End_Date_Division_Wide__c, Case_Number__r.Name, Case_Type__c
    FROM DEL_Placement__c
    WHERE PID__c IN {woppids}
    AND Service_Name__c != null
    AND Placement_Start_Date_Division_Wide__c != null
    AND Latest_Version_Placement__c = True
    ORDER BY Placement_Start_Date_Division_Wide__c DESC
    ''', woppids = woppids)             #Query for last placement info.

    wodf = soql_df(woq)          #Create DF

    wodf = wodf.sort_values(by = ['PID__c','Placement_Start_Date_Division_Wide__c'], ascending = [True, False]).copy()
    wodf = wodf.drop_duplicates('PID__c').copy()    #Sort by Placement Start Date and remove duplicates
    c = [
        'PID', 'Last Placement ID', 'Last Placing Division', 'Service', 'Last Placement Removal Reason','State', 'Placement Start Date','Placement End Date', 'Case Type',
         'Case Owner', 'Service Area','Supervisor', 'Case Number'
    ]

    wodf.columns = c                    #Change column names
    old_cols = [c for c in woplacement.columns if (c not in wodf.columns and c not in ('Placement ID', 'Division')) or c == 'PID']  # get columns not in new placement table.
    oldwoplacement = woplacement[old_cols].copy() #remove old placement columns.

    woplacement1 = oldwoplacement.merge(wodf, how='left', left_on = 'PID', right_on = 'PID')
    ord_columns = [
        'PID', 'Name', 'Birth Date', 'Age', 'Age (Years)', 'Custody Start Date','Last Placing Division','Service','Case Number', 'Case Type', 'Case Owner',
        'Supervisor','Service Area','State','Last Placement ID','Placement Start Date','Placement End Date','Who Has Custody','Runaway', 'MCI','MA Start Date', 'MA End Date', 'MA Code', 'Last Placement Removal Reason'
    ] 
    woplacement1 = woplacement1[ord_columns].copy() # Create new KIC without placement with last placement information.
    woplacement1.loc[woplacement1['MA End Date']== '2299-12-31', ['MA End Date']] = '2050-12-31'
    woplacement1.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date', 'Placement End Date','MA Start Date', 'MA End Date']]= woplacement1.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date', 'Placement End Date','MA Start Date', 'MA End Date']].apply(pd.to_datetime) 


    # Export to excel
    date_ = dt.date.today()
    month = date_.month
    day = date_.day
    year = date_.year
    location = fr'H:\Python_Programs\SF\Weekly\Results\KIC_report{month}.{day}.{year}.xlsx'
    writer = pd.ExcelWriter(location, engine = 'xlsxwriter', datetime_format = 'mm/dd/yyyy')
    wplacement.to_excel(writer, sheet_name = 'KIC W Active Placement', index=False)
    woplacement1.to_excel(writer, sheet_name = 'KIC WO Active Placement', index=False)
    writer.save()

    #Adjust column sizes
    book = load_workbook(location)
    for sheet in book.sheetnames:
        worksheet = book[sheet]
        column_size(worksheet)
    book.save(location)

    #Print Results
    r_loc = fr'H:\Python_Programs\SF\Weekly\Results\KIC_results{month}.{day}.{year}.txt'
    with open(r_loc, 'w') as file:
        kidswp = len(wplacement['PID'].unique())
        kidswop = len(woplacement1['PID'].unique())
        print(f'''Results:
        Kids in Custody with Active Placement: {kidswp}
        Kids in Custody without Active Placement: {kidswop}
        Total Kids in Custody: {kidswp + kidswop}''', file = file)

    return wplacement1

